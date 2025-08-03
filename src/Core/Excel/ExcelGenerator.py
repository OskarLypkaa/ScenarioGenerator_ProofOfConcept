import json
import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage
from utilities.logger import log

class ExcelGenerator:
    def __init__(self, jsonPath: str, templatePath: str, outputPath: str):
        self.jsonPath = jsonPath
        self.templatePath = templatePath
        self.outputPath = outputPath
        self.steps = []
        self.wb = None
        self.ws: Worksheet = None
        self.startRow = 4  
        self.subStepCounter = 1
        self.currentRow = self.startRow

    def generate(self):
        self._validatePaths()
        self._loadData()
        self._copyTemplate()
        self._loadTemplate()
        self._writeSteps()
        self._saveWorkbook()

    def _validatePaths(self):
        if not self.jsonPath or not os.path.exists(self.jsonPath):
            raise FileNotFoundError(f"❌ JSON file not found: {self.jsonPath}")
        if not self.templatePath or not os.path.exists(self.templatePath):
            raise FileNotFoundError(f"❌ Excel template not found: {self.templatePath}")
        if not self.templatePath.lower().endswith(".xlsm"):
            raise ValueError("❌ Template file must be .xlsm (macro-enabled)")

    def _loadData(self):
        with open(self.jsonPath, "r", encoding="utf-8") as f:
            self.steps = json.load(f)

    def _copyTemplate(self):
        try:
            shutil.copyfile(self.templatePath, self.outputPath)
            log.info(f"Template copied to: {self.outputPath}")
        except Exception as e:
            log.error(f"Failed to copy template: {e}")
            raise RuntimeError(f"Failed to copy template: {e}")

    def _loadTemplate(self):
        self.wb = load_workbook(self.outputPath, keep_vba=True)  # ← KLUCZOWE
        self.ws = self.wb.active

    def _writeSteps(self):
        for step in self.steps:
            self._writeRow(step)
            self.currentRow += 1
            self.subStepCounter += 1

    def _writeRow(self, step: dict):
        srcRow = self.startRow
        destRow = self.currentRow

        
        for col in range(1, 6):
            srcCell = self.ws.cell(row=srcRow, column=col)
            dstCell = self.ws.cell(row=destRow, column=col)

            dstCell.font = copy(srcCell.font)
            dstCell.fill = copy(srcCell.fill)
            dstCell.border = copy(srcCell.border)
            dstCell.alignment = copy(srcCell.alignment)
            dstCell.number_format = srcCell.number_format

        
        self.ws.row_dimensions[destRow].height = self.ws.row_dimensions[srcRow].height

       
        self.ws.cell(row=destRow, column=1).value = f"1.{self.subStepCounter}"  # A
        self.ws.cell(row=destRow, column=2).value = step.get("Taken Action", "")  # B
        self.ws.cell(row=destRow, column=4).value = step.get("Expected Result", "")  # D

        
        self._insertImage(step.get("Taken Action Picture"), f"C{destRow}")  # C
        self._insertImage(step.get("Expected Result Picture"), f"E{destRow}")  # E

    def _insertImage(self, imagePath: str, cellRef: str):
        if not imagePath or not os.path.exists(imagePath):
            log.warning(f"Image not found: {imagePath}")
            return
        try:
            excelImage = ExcelImage(imagePath)
            self.ws.add_image(excelImage, cellRef)
        except Exception as e:
            log.warning(f"Failed to insert image: {imagePath} ({e})")

    def _saveWorkbook(self):
        self.wb.save(self.outputPath)
        log.info(f"Excel macro-enabled file generated: {self.outputPath}")
